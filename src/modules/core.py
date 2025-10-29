import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import uuid 
import sys 

# Definir o caminho base do aplicativo para que o Excel seja encontrado
if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    _base_path = sys._MEIPASS
else:
    _base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))

EXCEL_FILE_NAME = 'financas_pessoais.xlsx'
EXCEL_DIR = 'data' 
EXCEL_FILE = os.path.join(_base_path, EXCEL_DIR, EXCEL_FILE_NAME)

DEFAULT_CATEGORIES_SHEET = 'Categorias'
DEFAULT_BUDGET_SHEET = 'OrcamentoMensal'
DEFAULT_LOANS_SHEET = 'Empréstimos' 
DEFAULT_DEBTS_SHEET = 'DívidasFuturas' 

class CoreManager:
    def __init__(self):
        self._ensure_excel_file_exists()

    def _ensure_excel_file_exists(self):
        """Verifica se o arquivo Excel e as abas padrão existem, criando-os se necessário."""
        excel_folder_path = os.path.dirname(EXCEL_FILE)
        if not os.path.exists(excel_folder_path):
            os.makedirs(excel_folder_path)
            print(f"Diretório '{excel_folder_path}' criado.")

        if not os.path.exists(EXCEL_FILE):
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
                pd.DataFrame(columns=['Categoria']).to_excel(writer, sheet_name=DEFAULT_CATEGORIES_SHEET, index=False)
                pd.DataFrame(columns=['MesAno', 'Categoria', 'Limite']).to_excel(writer, sheet_name=DEFAULT_BUDGET_SHEET, index=False)
                pd.DataFrame(columns=['ID', 'Tipo', 'ParteEnvolvida', 'ValorOriginal', 'Juros%', 'NumParcelas', 'ParcelasPagas', 'Status']).to_excel(writer, sheet_name=DEFAULT_LOANS_SHEET, index=False)
                pd.DataFrame(columns=['ID', 'Descricao', 'Valor', 'DataVencimento', 'Status', 'Recorrencia', 'RecorrenciaMeses', 'Categoria']).to_excel(writer, sheet_name=DEFAULT_DEBTS_SHEET, index=False)
            print(f"Arquivo Excel '{EXCEL_FILE}' e abas padrão criadas.")
        else:
            wb = load_workbook(EXCEL_FILE)
            if DEFAULT_CATEGORIES_SHEET not in wb.sheetnames:
                wb.create_sheet(DEFAULT_CATEGORIES_SHEET)
                wb[DEFAULT_CATEGORIES_SHEET].append(['Categoria'])
            if DEFAULT_BUDGET_SHEET not in wb.sheetnames:
                wb.create_sheet(DEFAULT_BUDGET_SHEET)
                wb[DEFAULT_BUDGET_SHEET].append(['MesAno', 'Categoria', 'Limite'])
            if DEFAULT_LOANS_SHEET not in wb.sheetnames:
                wb.create_sheet(DEFAULT_LOANS_SHEET)
                if wb[DEFAULT_LOANS_SHEET].max_row == 0 or 'ID' not in [cell.value for cell in wb[DEFAULT_LOANS_SHEET][1]]:
                    wb[DEFAULT_LOANS_SHEET].append(['ID', 'Tipo', 'ParteEnvolvida', 'ValorOriginal', 'Juros%', 'NumParcelas', 'ParcelasPagas', 'Status'])
            if DEFAULT_DEBTS_SHEET not in wb.sheetnames:
                ws = wb.create_sheet(DEFAULT_DEBTS_SHEET)
                ws.append(['ID', 'Descricao', 'Valor', 'DataVencimento', 'Status', 'Recorrencia', 'RecorrenciaMeses', 'Categoria'])
            wb.save(EXCEL_FILE)
            print(f"Verificação de abas padrão concluída em '{EXCEL_FILE}'.")

    def _get_sheet_name(self, month_year):
        return month_year

    def load_data(self, sheet_name):
        """Carrega os dados de uma aba específica do Excel para um DataFrame."""
        try:
            df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
            
            # Garante que 'ID' é string se existir
            if 'ID' in df.columns:
                df['ID'] = df['ID'].astype(str)
            
            # Garante colunas esperadas para abas mensais
            if sheet_name not in [DEFAULT_CATEGORIES_SHEET, DEFAULT_BUDGET_SHEET, DEFAULT_LOANS_SHEET, DEFAULT_DEBTS_SHEET]:
                expected_cols = ['ID', 'Data', 'Tipo', 'Descricao', 'Categoria', 'Valor', 'MeioPagamento'] 
                for col in expected_cols:
                    if col not in df.columns:
                        df[col] = "Conta" if col == 'MeioPagamento' else None 
            # Garante colunas esperadas para a aba de empréstimos
            elif sheet_name == DEFAULT_LOANS_SHEET:
                expected_cols = ['ID', 'Tipo', 'ParteEnvolvida', 'ValorOriginal', 'Juros%', 'NumParcelas', 'ParcelasPagas', 'Status']
                for col in expected_cols:
                    if col not in df.columns:
                        df[col] = None
            # Garante colunas esperadas para a aba de dívidas futuras
            elif sheet_name == DEFAULT_DEBTS_SHEET:
                expected_cols = ['ID', 'Descricao', 'Valor', 'DataVencimento', 'Status', 'Recorrencia', 'RecorrenciaMeses', 'Categoria']
                for col in expected_cols:
                    if col not in df.columns:
                        df[col] = None 

            return df
        except FileNotFoundError:
            print(f"Erro: Arquivo '{EXCEL_FILE}' não encontrado no caminho: {EXCEL_FILE}")
            return pd.DataFrame()
        except Exception as e:
            print(f"Erro ao carregar a aba '{sheet_name}': {e}")
            return pd.DataFrame()

    def save_data(self, df, sheet_name):
        """Salva um DataFrame em uma aba específica do Excel."""
        try:
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                if 'ID' in df.columns:
                    df['ID'] = df['ID'].astype(str)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Dados salvos na aba '{sheet_name}' com sucesso.")
            return True
        except Exception as e:
            print(f"Erro ao salvar dados na aba '{sheet_name}': {e}")
            return False

    def create_monthly_sheet(self, month_year):
        """Cria uma aba para o mês e ano especificados se não existir."""
        sheet_name = self._get_sheet_name(month_year)
        try:
            wb = load_workbook(EXCEL_FILE)
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(['ID', 'Data', 'Tipo', 'Descricao', 'Categoria', 'Valor', 'MeioPagamento'])
                wb.save(EXCEL_FILE)
                print(f"Aba '{sheet_name}' criada com sucesso.")
            return True
        except Exception as e:
            print(f"Erro ao criar aba '{sheet_name}': {e}")
            return False

    def get_monthly_transactions(self, month_year):
        """Retorna os lançamentos de um mês específico."""
        sheet_name = self._get_sheet_name(month_year)
        return self.load_data(sheet_name)

    def add_transaction(self, month_year, data):
        """Adiciona um novo lançamento à aba mensal."""
        sheet_name = self._get_sheet_name(month_year)
        self.create_monthly_sheet(month_year)
        current_df = self.load_data(sheet_name)

        if current_df.empty:
            data['ID'] = 1
        else:
            max_id_numeric = pd.to_numeric(current_df['ID'], errors='coerce').fillna(0).astype(int).max()
            data['ID'] = max_id_numeric + 1
        
        new_transaction_df = pd.DataFrame([data])
        updated_df = pd.concat([current_df, new_transaction_df], ignore_index=True)
        return self.save_data(updated_df, sheet_name)

    def update_transaction(self, month_year, transaction_id, new_data):
        """Atualiza um lançamento específico na aba mensal."""
        sheet_name = self._get_sheet_name(month_year) 
        df = self.load_data(sheet_name)
        if not df.empty:
            df['ID'] = df['ID'].astype(str)
            transaction_id = str(transaction_id)

            idx = df[df['ID'] == transaction_id].index
            if not idx.empty:
                for key, value in new_data.items():
                    df.loc[idx, key] = value
                return self.save_data(df, sheet_name)
        return False

    def delete_transaction(self, month_year, transaction_id):
        """Exclui um lançamento específico da aba mensal."""
        sheet_name = self._get_sheet_name(month_year)
        df = self.load_data(sheet_name)
        if not df.empty:
            df['ID'] = df['ID'].astype(str)
            transaction_id = str(transaction_id)
            
            df = df[df['ID'] != transaction_id]
            return self.save_data(df, sheet_name)
        return False

    def get_categories(self):
        """Retorna as categorias personalizadas."""
        return self.load_data(DEFAULT_CATEGORIES_SHEET)

    def add_category(self, category_name):
        """Adiciona uma nova categoria."""
        df = self.load_data(DEFAULT_CATEGORIES_SHEET)
        if df.empty or category_name not in df['Categoria'].values:
            new_row = pd.DataFrame([{'Categoria': category_name}])
            df = pd.concat([df, new_row], ignore_index=True)
            return self.save_data(df, DEFAULT_CATEGORIES_SHEET)
        return False

    def remove_category(self, category_name):
        """Remove uma categoria."""
        df = self.load_data(DEFAULT_CATEGORIES_SHEET)
        if not df.empty and category_name in df['Categoria'].values:
            df = df[df['Categoria'] != category_name]
            return self.save_data(df, DEFAULT_CATEGORIES_SHEET)
        return False

    def get_budgets(self):
        """Retorna os orçamentos mensais."""
        return self.load_data(DEFAULT_BUDGET_SHEET)

    def set_budget(self, month_year, category, limit):
        """Define ou atualiza o limite de orçamento para uma categoria em um mês."""
        df = self.load_data(DEFAULT_BUDGET_SHEET)
        
        updated = False
        for idx, row in df.iterrows():
            if str(row['MesAno']) == month_year and str(row['Categoria']) == category:
                df.loc[idx, 'Limite'] = limit
                updated = True
                break
        
        if not updated:
            new_row = pd.DataFrame([{'MesAno': month_year, 'Categoria': category, 'Limite': limit}])
            df = pd.concat([df, new_row], ignore_index=True)
            
        return self.save_data(df, DEFAULT_BUDGET_SHEET)

    def get_loans(self):
        """Retorna os empréstimos registrados."""
        return self.load_data(DEFAULT_LOANS_SHEET)

    def add_loan(self, data):
        """Adiciona um novo empréstimo/dívida. 'data' já deve conter o 'ID' (UUID)."""
        df = self.load_data(DEFAULT_LOANS_SHEET)
        
        new_loan_df = pd.DataFrame([data])
        df = pd.concat([df, new_loan_df], ignore_index=True)
        return self.save_data(df, DEFAULT_LOANS_SHEET)

    def update_loan(self, loan_id, new_data):
        """Atualiza um empréstimo específico."""
        df = self.load_data(DEFAULT_LOANS_SHEET)
        if not df.empty:
            df['ID'] = df['ID'].astype(str)
            loan_id = str(loan_id)

            idx = df[df['ID'] == loan_id].index
            if not idx.empty:
                for key, value in new_data.items():
                    df.loc[idx, key] = value
                return self.save_data(df, DEFAULT_LOANS_SHEET)
        return False

    def add_debt(self, data):
        df = self.load_data(DEFAULT_DEBTS_SHEET)
        data['ID'] = str(uuid.uuid4())
        new_debt_df = pd.DataFrame([data])
        df = pd.concat([df, new_debt_df], ignore_index=True)
        return self.save_data(df, DEFAULT_DEBTS_SHEET)

    def update_debt(self, debt_id, new_data):
        df = self.load_data(DEFAULT_DEBTS_SHEET)
        if not df.empty:
            df['ID'] = df['ID'].astype(str)
            debt_id = str(debt_id)
            idx = df[df['ID'] == debt_id].index
            if not idx.empty:
                for key, value in new_data.items():
                    df.loc[idx, key] = value
                return self.save_data(df, DEFAULT_DEBTS_SHEET)
        return False

    def delete_debt(self, debt_id):
        df = self.load_data(DEFAULT_DEBTS_SHEET)
        if not df.empty:
            df['ID'] = df['ID'].astype(str)
            debt_id = str(debt_id)
            df = df[df['ID'] != debt_id]
            return self.save_data(df, DEFAULT_DEBTS_SHEET)
        return False

    def get_debts(self):
        return self.load_data(DEFAULT_DEBTS_SHEET)