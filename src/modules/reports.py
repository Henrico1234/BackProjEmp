import pandas as pd
from datetime import datetime
import openpyxl 
import os 
from fpdf import FPDF
import sqlite3

from .core import CoreManager, DB_FILE
from .monthly_control import MonthlyControlManager

class ReportManager:
    def __init__(self, core_manager: CoreManager, monthly_control_manager: MonthlyControlManager):
        self.core = core_manager
        self.monthly_control = monthly_control_manager 

    def _get_all_transaction_sheets(self):
        if not os.path.exists(DB_FILE): 
            print(f"Aviso: Arquivo de banco de dados '{DB_FILE}' não encontrado.")
            return []
            
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT MesAno FROM Transacoes;")
                monthly_sheets = [item[0] for item in cursor.fetchall()]
                return monthly_sheets
        except Exception as e:
            print(f"Erro ao listar 'MesAno' da tabela Transacoes: {e}")
            return []

    def get_all_transactions_in_period(self, start_date: datetime, end_date: datetime):

        all_transactions_df = pd.DataFrame()
        
        all_sheets = self._get_all_transaction_sheets() 

        start_ts = pd.Timestamp(start_date)
        end_ts = pd.Timestamp(end_date)

        for sheet_name in all_sheets:
            try:
                sheet_month_year_dt = datetime.strptime(sheet_name, "%m-%Y")
                
                sheet_start_date_dt = sheet_month_year_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                
                next_month_dt = (pd.Timestamp(sheet_month_year_dt) + pd.DateOffset(months=1)).to_pydatetime()
                sheet_end_date_dt = (next_month_dt - pd.DateOffset(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)

                sheet_start_ts = pd.Timestamp(sheet_start_date_dt)
                sheet_end_ts = pd.Timestamp(sheet_end_date_dt)

                if not (sheet_end_ts < start_ts or sheet_start_ts > end_ts):
                
                    month_df = self.core.get_monthly_transactions(sheet_name) 
                    
                    if not month_df.empty:
                        if 'Data' in month_df.columns:
                            month_df['Data'] = pd.to_datetime(month_df['Data'], errors='coerce')
                        else:
                            print(f"Aviso: Aba '{sheet_name}' não possui coluna 'Data'. Ignorando.")
                            continue

                        if 'Valor' in month_df.columns:
                            month_df['Valor'] = pd.to_numeric(month_df['Valor'], errors='coerce')
                        else:
                            print(f"Aviso: Aba '{sheet_name}' não possui coluna 'Valor'. Ignorando.")
                            continue
                        
                        expected_cols = ['ID', 'Data', 'Tipo', 'Descricao', 'Categoria', 'Valor', 'MeioPagamento']
                        for col in expected_cols:
                            if col not in month_df.columns:
                                month_df[col] = "Conta" if col == 'MeioPagamento' else None 

                        month_df_filtered = month_df[(month_df['Data'] >= start_ts) & (month_df['Data'] <= end_ts)]
                        all_transactions_df = pd.concat([all_transactions_df, month_df_filtered[expected_cols]], ignore_index=True)
            
            except ValueError as ve:
                print(f"Aviso: Problema de formato de data na aba '{sheet_name}' ou nome de aba inválido: {ve}")
                continue 
            except Exception as e:
                print(f"Erro ao processar aba '{sheet_name}' para relatório: {e}")
                continue
        
        if not all_transactions_df.empty and 'Data' in all_transactions_df.columns:
            return all_transactions_df.dropna(subset=['Data'])
        return all_transactions_df

    def generate_financial_summary(self, start_date: datetime, end_date: datetime, category: str = None):

        transactions_df = self.get_all_transactions_in_period(start_date, end_date)
        
        if transactions_df.empty:
            return {
                'Ganhos Totais': 0.0,
                'Despesas Totais': 0.0,
                'Saldo Total': 0.0,
                'Despesas por Categoria': pd.Series(dtype=float),
                'Ganhos por Categoria': pd.Series(dtype=float)
            }

        if category and category.lower() != "todas":
            transactions_df = transactions_df[transactions_df['Categoria'].astype(str).str.lower() == category.lower()]
            if transactions_df.empty: 
                return {
                    'Ganhos Totais': 0.0,
                    'Despesas Totais': 0.0,
                    'Saldo Total': 0.0,
                    'Despesas por Categoria': pd.Series(dtype=float),
                    'Ganhos por Categoria': pd.Series(dtype=float)
                }

        gains_df = transactions_df[transactions_df['Tipo'].astype(str).str.lower() == 'ganho']
        expenses_df = transactions_df[transactions_df['Tipo'].astype(str).str.lower() == 'despesa']

        total_gains = gains_df['Valor'].sum()
        total_expenses = expenses_df['Valor'].sum()
        total_balance = total_gains - total_expenses

        expenses_by_category = expenses_df.groupby(expenses_df['Categoria'].astype(str))['Valor'].sum().sort_values(ascending=False)
        gains_by_category = gains_df.groupby(gains_df['Categoria'].astype(str))['Valor'].sum().sort_values(ascending=False)

        return {
            'Ganhos Totais': total_gains,
            'Despesas Totais': total_expenses,
            'Saldo Total': total_balance,
            'Despesas por Categoria': expenses_by_category,
            'Ganhos por Categoria': gains_by_category
        }
    
    def export_summary_to_csv(self, summary_data: dict, filename="relatorio_financeiro.csv"):
        """Exporta o resumo financeiro para um arquivo CSV.
        (Este método não precisa de alterações)
        """
        try:
            if not os.path.exists('data'):
                os.makedirs('data')
            
            filepath = os.path.join('data', filename)

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(f"Resumo Financeiro - Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                f.write(f"Ganhos Totais: R${summary_data['Ganhos Totais']:.2f}\n")
                f.write(f"Despesas Totais: R${summary_data['Despesas Totais']:.2f}\n")
                f.write(f"Saldo Total: R${summary_data['Saldo Total']:.2f}\n\n")
                
                f.write("Despesas por Categoria:\n")
                if not summary_data['Despesas por Categoria'].empty:
                    for cat, val in summary_data['Despesas por Categoria'].items():
                        f.write(f"- {cat}: R${val:.2f}\n")
                else:
                    f.write("Nenhuma despesa para exibir.\n")
                f.write("\n")

                f.write("Ganhos por Categoria:\n")
                if not summary_data['Ganhos por Categoria'].empty:
                    for cat, val in summary_data['Ganhos por Categoria'].items():
                        f.write(f"- {cat}: R${val:.2f}\n")
                else:
                    f.write("Nenhum ganho para exibir.\n")
                f.write("\n")

            print(f"Relatório exportado para {filepath}")
            return True
        except Exception as e:
            print(f"Erro ao exportar relatório para CSV: {e}")
            return False

    def export_summary_to_pdf(self, summary_data: dict, filename="relatorio_financeiro.pdf"):
        """Exporta o resumo financeiro para um arquivo PDF.
        (Este método foi corrigido)
        """
        try:
            if not os.path.exists('data'):
                os.makedirs('data')
            
            filepath = os.path.join('data', filename)

            pdf = FPDF()
            pdf.add_page()
            
            font_family = "Arial"
            
            try:
                pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
                pdf.set_font("DejaVu", "", 12)
                font_family = "DejaVu"
            except Exception:
                print("Aviso: Fonte DejaVuSans não encontrada. Usando Arial (pode ter problemas com R$).")
                pdf.set_font("Arial", "", 12)
                font_family = "Arial"

            
            # Título
            pdf.set_font(font_family, size=16)
            pdf.cell(0, 10, "Resumo Financeiro", 0, 1, "C")
            pdf.set_font(font_family, size=10)
            pdf.cell(0, 5, f"Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 1, "C")
            pdf.ln(10)

            pdf.set_font(font_family, size=12, style='B')
            pdf.cell(0, 8, "Totais:", 0, 1, "L")
            pdf.set_font(font_family, size=12)
            pdf.set_text_color(0, 128, 0)
            pdf.cell(0, 7, f"Ganhos Totais: R$ {summary_data['Ganhos Totais']:.2f}".replace('.', ','), 0, 1, "L")
            pdf.set_text_color(255, 0, 0)
            pdf.cell(0, 7, f"Despesas Totais: R$ {summary_data['Despesas Totais']:.2f}".replace('.', ','), 0, 1, "L")
            
            balance_color = (0, 0, 0)
            if summary_data['Saldo Total'] > 0:
                balance_color = (0, 128, 0)
            elif summary_data['Saldo Total'] < 0:
                balance_color = (255, 0, 0)
            pdf.set_text_color(*balance_color)
            pdf.set_font(font_family, size=13, style='B')
            pdf.cell(0, 10, f"Saldo Total: R$ {summary_data['Saldo Total']:.2f}".replace('.', ','), 0, 1, "L")
            pdf.ln(5)
            pdf.set_text_color(0, 0, 0)

            pdf.set_font(font_family, size=12, style='B')
            pdf.cell(0, 8, "Despesas por Categoria:", 0, 1, "L")
            pdf.set_font(font_family, size=11)
            if not summary_data['Despesas por Categoria'].empty:
                for cat, val in summary_data['Despesas por Categoria'].items():
                    pdf.cell(0, 7, f"- {cat}: R$ {val:.2f}".replace('.', ','), 0, 1, "L")
            else:
                pdf.cell(0, 7, "Nenhuma despesa para exibir.", 0, 1, "L")
            pdf.ln(5)

            pdf.set_font(font_family, size=12, style='B')
            pdf.cell(0, 8, "Ganhos por Categoria:", 0, 1, "L")
            pdf.set_font(font_family, size=11)
            if not summary_data['Ganhos por Categoria'].empty:
                for cat, val in summary_data['Ganhos por Categoria'].items():
                    pdf.cell(0, 7, f"- {cat}: R$ {val:.2f}".replace('.', ','), 0, 1, "L")
            else:
                pdf.cell(0, 7, "Nenhum ganho para exibir.", 0, 1, "L")
            pdf.ln(5)

            pdf.output(filepath)
            print(f"Relatório exportado para PDF em: {filepath}")
            return True
        except Exception as e:
            if "FPDF error" in str(e) and "Could not include font" in str(e):
                print("--- ERRO DE FONTE FPDF ---")
                print("Ocorreu um erro ao gerar o PDF, provavelmente a fonte DejaVuSans não foi encontrada.")
                print("Tenta instalar a fonte no teu sistema (ex: sudo apt-get install fonts-dejavu-core) ou ignora o PDF por agora.")
                return False
            print(f"Erro ao exportar relatório para PDF: {e}")
            return False